"""
Data validation utilities
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Tuple
import re
from pathlib import Path

class DataValidator:
    """Data validation utility class"""
    
    def __init__(self):
        self.validation_rules = {}
        self.custom_validators = {}
    
    def validate_file_format(self, file_path: str) -> Dict[str, Any]:
        """Validate if file format is supported"""
        result = {
            'is_valid': False,
            'file_type': None,
            'errors': [],
            'warnings': []
        }
        
        try:
            file_path = Path(file_path)
            
            if not file_path.exists():
                result['errors'].append(f"File does not exist: {file_path}")
                return result
            
            if not file_path.is_file():
                result['errors'].append(f"Path is not a file: {file_path}")
                return result
            
            # Check file extension
            ext = file_path.suffix.lower()
            supported_formats = {'.csv': 'CSV', '.xlsx': 'Excel', '.xls': 'Excel'}
            
            if ext not in supported_formats:
                result['errors'].append(f"Unsupported file format: {ext}")
                return result
            
            result['file_type'] = supported_formats[ext]
            
            # Check file size (warn if very large)
            file_size = file_path.stat().st_size
            if file_size > 100 * 1024 * 1024:  # 100MB
                result['warnings'].append(f"Large file size: {file_size / (1024*1024):.1f} MB")
            
            # Try to peek at the file content
            if ext == '.csv':
                self._validate_csv_content(file_path, result)
            elif ext in ['.xlsx', '.xls']:
                self._validate_excel_content(file_path, result)
            
            result['is_valid'] = len(result['errors']) == 0
            
        except Exception as e:
            result['errors'].append(f"Validation error: {str(e)}")
        
        return result
    
    def _validate_csv_content(self, file_path: Path, result: Dict):
        """Validate CSV file content"""
        try:
            # Try to read first few lines
            with open(file_path, 'r', encoding='utf-8') as f:
                first_lines = [f.readline() for _ in range(5)]
            
            # Check if file is empty
            if not any(line.strip() for line in first_lines):
                result['errors'].append("CSV file appears to be empty")
                return
            
            # Check for consistent number of columns
            delimiters = [',', ';', '\t', '|']
            column_counts = {}
            
            for delimiter in delimiters:
                counts = [line.count(delimiter) for line in first_lines if line.strip()]
                if counts and len(set(counts)) <= 2:  # Allow some variation
                    column_counts[delimiter] = counts[0]
            
            if not column_counts:
                result['warnings'].append("Could not detect consistent delimiter")
            
        except UnicodeDecodeError:
            result['errors'].append("CSV file encoding issue - try different encoding")
        except Exception as e:
            result['warnings'].append(f"CSV validation warning: {str(e)}")
    
    def _validate_excel_content(self, file_path: Path, result: Dict):
        """Validate Excel file content"""
        try:
            # Try to read Excel file metadata
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            
            if not workbook.sheetnames:
                result['errors'].append("Excel file has no worksheets")
                return
            
            # Check first worksheet
            worksheet = workbook[workbook.sheetnames[0]]
            
            if worksheet.max_row == 1 and worksheet.max_column == 1:
                cell_value = worksheet.cell(1, 1).value
                if not cell_value:
                    result['warnings'].append("Excel file appears to be empty")
            
            workbook.close()
            
        except Exception as e:
            result['warnings'].append(f"Excel validation warning: {str(e)}")
    
    def validate_dataframe(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Comprehensive DataFrame validation"""
        result = {
            'is_valid': True,
            'errors': [],
            'warnings': [],
            'info': [],
            'statistics': {}
        }
        
        try:
            # Basic structure validation
            if df.empty:
                result['errors'].append("DataFrame is empty")
                result['is_valid'] = False
                return result
            
            if len(df.columns) == 0:
                result['errors'].append("DataFrame has no columns")
                result['is_valid'] = False
                return result
            
            # Column name validation
            self._validate_column_names(df, result)
            
            # Data type validation
            self._validate_data_types(df, result)
            
            # Missing data validation
            self._validate_missing_data(df, result)
            
            # Duplicate validation
            self._validate_duplicates(df, result)
            
            # Value range validation
            self._validate_value_ranges(df, result)
            
            # Generate statistics
            result['statistics'] = self._generate_validation_statistics(df)
            
        except Exception as e:
            result['errors'].append(f"Validation error: {str(e)}")
            result['is_valid'] = False
        
        return result
    
    def _validate_column_names(self, df: pd.DataFrame, result: Dict):
        """Validate column names"""
        # Check for duplicate column names
        if len(df.columns) != len(set(df.columns)):
            duplicates = [col for col in df.columns if list(df.columns).count(col) > 1]
            result['errors'].append(f"Duplicate column names found: {set(duplicates)}")
        
        # Check for empty column names
        empty_cols = [i for i, col in enumerate(df.columns) if not str(col).strip()]
        if empty_cols:
            result['warnings'].append(f"Empty column names at positions: {empty_cols}")
        
        # Check for problematic characters in column names
        problematic_cols = []
        for col in df.columns:
            col_str = str(col)
            if re.search(r'[^\w\s\-_.]', col_str):
                problematic_cols.append(col)
        
        if problematic_cols:
            result['warnings'].append(f"Columns with special characters: {problematic_cols}")
    
    def _validate_data_types(self, df: pd.DataFrame, result: Dict):
        """Validate data types"""
        # Check for mixed types in columns
        mixed_type_cols = []
        
        for col in df.columns:
            if df[col].dtype == 'object':
                # Check if column contains mixed numeric and text data
                non_null_data = df[col].dropna()
                if len(non_null_data) > 0:
                    numeric_count = 0
                    for value in non_null_data.head(100):  # Sample first 100
                        try:
                            float(value)
                            numeric_count += 1
                        except (ValueError, TypeError):
                            pass
                    
                    if 0 < numeric_count < len(non_null_data.head(100)):
                        mixed_type_cols.append(col)
        
        if mixed_type_cols:
            result['warnings'].append(f"Columns with mixed data types: {mixed_type_cols}")
    
    def _validate_missing_data(self, df: pd.DataFrame, result: Dict):
        """Validate missing data patterns"""
        missing_stats = df.isnull().sum()
        
        # Check for columns with excessive missing data
        high_missing_cols = []
        for col, missing_count in missing_stats.items():
            missing_pct = (missing_count / len(df)) * 100
            if missing_pct > 90:
                high_missing_cols.append((col, missing_pct))
        
        if high_missing_cols:
            result['warnings'].append(f"Columns with >90% missing data: {high_missing_cols}")
        
        # Check for completely empty rows
        empty_rows = df.isnull().all(axis=1).sum()
        if empty_rows > 0:
            result['warnings'].append(f"Found {empty_rows} completely empty rows")
    
    def _validate_duplicates(self, df: pd.DataFrame, result: Dict):
        """Validate duplicate data"""
        duplicate_count = df.duplicated().sum()
        
        if duplicate_count > 0:
            duplicate_pct = (duplicate_count / len(df)) * 100
            if duplicate_pct > 50:
                result['errors'].append(f"High duplicate rate: {duplicate_pct:.1f}% ({duplicate_count} rows)")
            elif duplicate_pct > 10:
                result['warnings'].append(f"Moderate duplicate rate: {duplicate_pct:.1f}% ({duplicate_count} rows)")
            else:
                result['info'].append(f"Found {duplicate_count} duplicate rows ({duplicate_pct:.1f}%)")
    
    def _validate_value_ranges(self, df: pd.DataFrame, result: Dict):
        """Validate value ranges for numeric columns"""
        range_issues = []
        
        for col in df.select_dtypes(include=[np.number]).columns:
            col_data = df[col].dropna()
            
            if len(col_data) == 0:
                continue
            
            # Check for infinite values
            if np.isinf(col_data).any():
                range_issues.append(f"{col}: Contains infinite values")
            
            # Check for extremely large values
            if col_data.abs().max() > 1e10:
                range_issues.append(f"{col}: Contains extremely large values")
            
            # Domain-specific validations
            col_lower = col.lower()
            
            if 'age' in col_lower:
                if (col_data < 0).any() or (col_data > 150).any():
                    range_issues.append(f"{col}: Unrealistic age values")
            
            elif 'percentage' in col_lower or 'percent' in col_lower:
                if (col_data < 0).any() or (col_data > 100).any():
                    range_issues.append(f"{col}: Percentage values outside 0-100 range")
            
            elif any(term in col_lower for term in ['price', 'cost', 'amount', 'salary']):
                if (col_data < 0).any():
                    negative_count = (col_data < 0).sum()
                    range_issues.append(f"{col}: Contains {negative_count} negative monetary values")
        
        if range_issues:
            result['warnings'].extend(range_issues)
    
    def _generate_validation_statistics(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Generate validation statistics"""
        return {
            'total_rows': len(df),
            'total_columns': len(df.columns),
            'numeric_columns': len(df.select_dtypes(include=[np.number]).columns),
            'categorical_columns': len(df.select_dtypes(include=['object']).columns),
            'datetime_columns': len(df.select_dtypes(include=['datetime']).columns),
            'total_missing_values': df.isnull().sum().sum(),
            'missing_percentage': (df.isnull().sum().sum() / (len(df) * len(df.columns))) * 100,
            'duplicate_rows': df.duplicated().sum(),
            'memory_usage_mb': df.memory_usage(deep=True).sum() / (1024 * 1024)
        }
    
    def validate_column_for_analysis(self, df: pd.DataFrame, column: str, analysis_type: str) -> Dict[str, Any]:
        """Validate if a column is suitable for specific analysis"""
        result = {
            'is_suitable': False,
            'issues': [],
            'recommendations': []
        }
        
        if column not in df.columns:
            result['issues'].append(f"Column '{column}' not found in dataframe")
            return result
        
        col_data = df[column]
        
        if analysis_type == 'correlation':
            if not pd.api.types.is_numeric_dtype(col_data):
                result['issues'].append("Column must be numeric for correlation analysis")
            elif col_data.nunique() < 2:
                result['issues'].append("Column has insufficient variation for correlation")
            elif col_data.isnull().sum() / len(col_data) > 0.5:
                result['issues'].append("Column has too many missing values (>50%)")
            else:
                result['is_suitable'] = True
        
        elif analysis_type == 'categorical':
            if col_data.nunique() > 50:
                result['issues'].append("Column has too many unique values for categorical analysis")
                result['recommendations'].append("Consider grouping values or using different analysis")
            elif col_data.nunique() < 2:
                result['issues'].append("Column has insufficient categories")
            else:
                result['is_suitable'] = True
        
        elif analysis_type == 'outlier_detection':
            if not pd.api.types.is_numeric_dtype(col_data):
                result['issues'].append("Column must be numeric for outlier detection")
            elif col_data.std() == 0:
                result['issues'].append("Column has no variation - cannot detect outliers")
            else:
                result['is_suitable'] = True
        
        elif analysis_type == 'time_series':
            if not pd.api.types.is_datetime64_any_dtype(col_data):
                try:
                    pd.to_datetime(col_data, errors='raise')
                    result['recommendations'].append("Column can be converted to datetime")
                    result['is_suitable'] = True
                except:
                    result['issues'].append("Column cannot be converted to datetime")
            else:
                result['is_suitable'] = True
        
        return result
    
    def suggest_data_improvements(self, df: pd.DataFrame) -> Dict[str, List[str]]:
        """Suggest improvements for data quality"""
        suggestions = {
            'critical': [],
            'recommended': [],
            'optional': []
        }
        
        # Missing data suggestions
        missing_data = df.isnull().sum()
        high_missing = missing_data[missing_data / len(df) > 0.3]
        
        if len(high_missing) > 0:
            suggestions['critical'].append(f"Handle missing data in columns: {list(high_missing.index)}")
        
        # Duplicate suggestions
        if df.duplicated().sum() > 0:
            suggestions['recommended'].append("Remove duplicate rows")
        
        # Data type suggestions
        for col in df.select_dtypes(include=['object']).columns:
            # Check if can be converted to numeric
            try:
                pd.to_numeric(df[col], errors='raise')
                suggestions['optional'].append(f"Convert '{col}' to numeric type")
            except:
                pass
            
            # Check if can be converted to datetime
            try:
                pd.to_datetime(df[col], errors='raise')
                suggestions['optional'].append(f"Convert '{col}' to datetime type")
            except:
                pass
        
        # Outlier suggestions
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        for col in numeric_cols:
            Q1 = df[col].quantile(0.25)
            Q3 = df[col].quantile(0.75)
            IQR = Q3 - Q1
            outliers = ((df[col] < (Q1 - 1.5 * IQR)) | (df[col] > (Q3 + 1.5 * IQR))).sum()
            
            if outliers > len(df) * 0.1:  # More than 10% outliers
                suggestions['recommended'].append(f"Investigate outliers in '{col}' ({outliers} values)")
        
        # Column name suggestions
        problematic_names = [col for col in df.columns if not str(col).replace('_', '').replace('-', '').isalnum()]
        if problematic_names:
            suggestions['optional'].append("Standardize column names")
        
        return suggestions